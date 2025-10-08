import pandas as pd
import numpy as np
import statsmodels.tsa.stattools as ts
import openpyxl
from tabulate import tabulate

excel_dataframe= openpyxl.load_workbook("model.xlsx")

dataframe=excel_dataframe.active

data=[]

for row in range(2, dataframe.max_row): #recuento de los numeros de filas 
    _row = [row,]
    for col in dataframe.iter_cols(2,dataframe.max_column):
        _row.append(col[row].value)

    data.append(_row)
    
headers = ["S&P", "PBI", "TCRM", "TIR", "IPC", "Empleo"] 
headers_align = (("center",)*6)

print(tabulate(data,headers=headers,tablefmt="fancy_grid", colalign=headers_align))



